"""
    xlsxファイルのシートを別のブックのシートにコビーする
    使い方：python3 copy_sheet.py wb1 sheet1 wb2 sheet2
    2019/01/10 Shimizu
"""
import openpyxl, sys

#****************workbook名からworkbookオブジェクトを取得する関数get_workbook
def get_workbook(workbook_name):
    try:
        workbookObj = openpyxl.load_workbook(workbook_name)
    except FileNotFoundError as e:
        print(e)
        exit()
    else:
        return workbookObj

#****************workbookオブジェクトとworksheet名からworksheetオブジェクトを取得する関数get_worksheet
def get_worksheet(workbookObj,worksheet_name):
    try:
        worksheetObj = workbookObj.get_sheet_by_name(worksheet_name)
    except KeyError as e:
        print(e)
        exit()
    else:
        return worksheetObj

#*******************コマンドライン引数の数が正しいか判定********************
if len(sys.argv) != 5:
    print("使い方：python3 copy_sheet.py wb1 sheet1 wb2 sheet2")
    print("コマンドライン引数は4つ:book,sheet,book,sheetの順")
    sys.exit()

#**************************変数宣言・代入***********************************
args = sys.argv
original_wb_name = args[1]
original_sheet_name = args[2]
copy_wb_name = args[3]
copy_sheet_name = args[4]

original_wb = get_workbook(original_wb_name)
original_sheet = get_worksheet(original_wb,original_sheet_name)
copy_wb = get_workbook(copy_wb_name)
copy_sheet = get_worksheet(copy_wb,copy_sheet_name)

#**************************メイン処理***************************************
for row_num in range(1,original_sheet.max_row + 1):
    for column_num in range(1,original_sheet.max_column +1):
        copy_sheet.cell(row=row_num,column=column_num).value = original_sheet.cell(row=row_num,column=column_num).value

copy_wb.save(copy_wb_name)


